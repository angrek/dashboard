#!/home/wrehfiel/ENV/bin/python2.7
#########################################################################
#
# Testing reading a workbook into a Django model to map out applications
# and application owners to actual servers.
#
# Boomer Rehfield - 07/23/2015
#
#########################################################################

import os
import re
from subprocess import call
from ssh import SSHClient

#from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

from django.utils import timezone
from server.models import AIXServer, LinuxServer, Power7Inventory, Storage

#these are need in django 1.7 and needed vs the django settings command
import django
from dashboard import settings
django.setup()


#s = Style(font=Font(name='Calibri', size=11, bold=True))



def get_worksheet_data():
    wb = load_workbook(filename = '/home/wrehfiel/ENV/dashboard/files/applist.xlsx', data_only=True)
    sheet_ranges = wb['1']
    print (sheet_ranges['D5'].value)
    print (sheet_ranges['D2'].value)
    print (sheet_ranges['A7'].value)
    print (sheet_ranges['F5'].value)

    columns = ('A', 'B', 'C', 'D', 'E', 'F', 'G')
    for num in range(1, 10):
        for column in columns:
            x = column + str(num)
            print (sheet_ranges[x].value)
#ws1 = wb.active




#ws1.title = 'AIX'
#ws2 = wb.create_sheet()
#ws2.title = 'Linux'



#    #need to put the timestamp in the filename
#    now = timezone.now()
#    timestamp = now.strftime('%m-%d-%Y')
#    filename = 'Unix_Configuration_Report' + timestamp + '.xlsx'
#    wb.save(filename)


    #c and f are just arbitrary vars right now for page one and two :\
#    c.style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
#    f.style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
#
#    command = 'echo "Unix Configuration Report" | mutt -a "' + filename + '" -s "Unix Configuration Report" -- boomer@wellcare.com'
#    os.system(command)







#start execution
if __name__ == '__main__':
    print "Reading the spreadsheet."
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dashboard.settings')
    get_worksheet_data()

    #print "Elapsed time: " + str(elapsed_time)







