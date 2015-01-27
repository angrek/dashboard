#!/home/wrehfiel/ENV/bin/python2.7
#########################################################################
#
# This is to pull data from Django and drop it into Excel for
# capacity planning
#
# Boomer Rehfield - 10/22/2014
#
#########################################################################

import os
import re
from subprocess import call
from ssh import SSHClient
from openpyxl import Workbook
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

from django.utils import timezone
from server.models import AIXServer, LinuxServer, Power7Inventory, Storage

#these are need in django 1.7 and needed vs the django settings command
import django
from dashboard import settings
django.setup()


s = Style(font=Font(name='Calibri', size=11, bold=True))



wb = Workbook()
ws1 = wb.active
ws1.title = 'AIX'
ws2 = wb.create_sheet()
ws2.title = 'Linux'


#create le pretty headers
ws1['A1'] = 'Host Name'
ws1['B1'] = 'OS'
ws1['C1'] = 'Physical/VM'
ws1['D1'] = 'IP'
ws1['E1'] = 'Mem(MB)'
ws1['F1'] = 'Database Name'
ws1['G1'] = 'Storage'
ws1['H1'] = 'CPU Cores'

ws2['A1'] = 'Host Name'
ws2['B1'] = 'OS'
ws2['C1'] = 'Physical/VM'
ws2['D1'] = 'IP'
ws2['E1'] = 'Mem(MB)'
ws2['F1'] = 'Database Name'
ws2['G1'] = 'Storage'
ws2['H1'] = 'CPU Cores'

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 10
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["F"].width = 20
ws1.column_dimensions["H"].width = 12

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["F"].width = 20
ws2.column_dimensions["H"].width = 12

list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

#to hell with writing all this out because the row dimensions are working...
for letter in list:
    cell = letter + '1'
    c = ws1[cell]
    f = ws2[cell]
    c.style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    f.style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

d = ws1.row_dimensions[1].height = 20
g = ws2.row_dimensions[1].height = 20

def get_server_data():
#starting line = 3
    line = 3
    counter = 0
    server_list = AIXServer.objects.filter(active=True, decommissioned=False)
    #shortening this just to speed up testing
    #server_list = AIXServer.objects.filter(name__contains='vio')
    for server in server_list:

        counter = counter + 1

        t = AIXServer.objects.filter(name=server, active=True)

        try:
            r = Power7Inventory.objects.get(name=server)
        except:
            pass

        try:
            p = Storage.objects.get(name=server)
        except:
            pass

        print str(counter) + ',' + str(server) + ',AIX,VM,' + server.ip_address.rstrip() + ',' + str(r.curr_mem) + ',,' + str(p.size) + ',' + str(r.curr_procs)

        cell = 'A' + str(line)
        ws1[cell] = str(server)

        cell = 'B' + str(line)
        ws1[cell] = 'AIX'

        cell = 'C' + str(line)
        ws1[cell] = 'VM'

        cell = 'D' + str(line)
        ws1[cell] = server.ip_address.rstrip()

        cell = 'E' + str(line)
        ws1[cell] = r.curr_mem

        cell = 'F' + str(line)
        ws1[cell] = ''

        cell = 'G' + str(line)
        ws1[cell] = p.size

        cell = 'H' + str(line)
        ws1[cell] = r.curr_procs



        #AIXServer.objects.filter(name=server).update(active=False, modified=timezone.now())
        line += 1
        aix_line = line


    #################STARTING LINUX SECTION############################
    line = 3
    counter = 0
    linux_server_list = LinuxServer.objects.all()
    #shortening this just to speed up testing
    #server_list = AIXServer.objects.filter(name__contains='vio')
    for server in linux_server_list:

        counter = counter + 1

        t = LinuxServer.objects.get(name=server)
        if t.active == False:
            #we don't care about inactive servers for capacity planning
            continue

        #try:
        #    r = Power7Inventory.objects.get(name=server)
        #except:
        #    pass
        #
        #try:
        #    p = Storage.objects.get(name=server)
        #except:
        #    pass

        print str(counter) + ',' + str(server) + ',Linux,VM,' + t.ip_address.rstrip() + ',' + str(t.memory) + ',,' + ' n/a' + ',' + str(t.cpu)

        cell = 'A' + str(line)
        ws2[cell] = str(server)

        cell = 'B' + str(line)
        ws2[cell] = 'Linux'

        cell = 'C' + str(line)
        ws2[cell] = 'VM'

        cell = 'D' + str(line)
        ws2[cell] = t.ip_address.rstrip()

        cell = 'E' + str(line)
        ws2[cell] = t.memory

        cell = 'F' + str(line)
        ws2[cell] = ''

        cell = 'G' + str(line)
        ws2[cell] = ''

        cell = 'H' + str(line)
        ws2[cell] = t.cpu



        #AIXServer.objects.filter(name=server).update(active=False, modified=timezone.now())
        line += 1



    #total
    cell = 'A' + str(line + 1)
    aix_cell = 'A' +str(aix_line + 1)
    ws1[aix_cell] = 'Total'
    ws2[cell] = 'Total'
    ws1[aix_cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    ws2[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

    cell = 'E' + str(line + 1)
    aix_cell = 'E' + str(aix_line + 1)
    sum = "=SUM(E3:E" + str((line - 1)) + ")"
    aix_sum = "=SUM(E3:E" + str((aix_line -1)) + ")"
    print aix_sum
    print sum
    ws1[aix_cell] = aix_sum
    ws2[cell] = sum
    ws1[aix_cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    ws2[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

    cell = 'G' + str(line + 1)
    aix_cell = 'G' + str(aix_line + 1)
    sum = "=SUM(G3:G" + str((line - 1)) + ")"
    aix_sum = "=SUM(G3:G" + str((aix_line - 1)) + ")"
    print aix_cell
    print aix_sum
    ws1[aix_cell] = aix_sum
    ws2[cell] = sum
    ws1[aix_cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    ws2[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

    cell = 'H' + str(line + 1)
    aix_cell = 'H' + str(aix_line + 1)
    sum = "=SUM(H3:H" + str((line - 1)) + ")"
    aix_sum = "=SUM(H3:H" + str((aix_line - 1)) + ")"
    ws1[aix_cell] = aix_sum
    ws2[cell] = sum
    ws1[aix_cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    ws2[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

#create le pretty headers
    cell = 'A' + str(aix_line + 3)
    ws1[cell] = 'Host Name'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'B' + str(aix_line + 3)
    ws1[cell] = 'OS'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'C' + str(aix_line + 3)
    ws1[cell] = 'Physical/VM'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'D' + str(aix_line + 3)
    ws1[cell] = 'IP'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'E' + str(aix_line + 3)
    ws1[cell] = 'Mem(MB)'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'F' + str(aix_line + 3)
    ws1[cell] = 'Database Name'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'G' + str(aix_line + 3)
    ws1[cell] = 'Storage'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    cell = 'H' + str(aix_line + 3)
    ws1[cell] = 'CPU Cores'
    ws1[cell].style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

    #ws2['A1'] = 'Host Name'
    #ws2['B1'] = 'OS'
    #ws2['C1'] = 'Physical/VM'
    #ws2['D1'] = 'IP'
    #ws2['E1'] = 'Mem(MB)'
    #ws2['F1'] = 'Database Name'
    #ws2['G1'] = 'Storage'
    #ws2['H1'] = 'CPU Cores'

    #need to put the timestamp in the filename
    now = timezone.now()
    timestamp = now.strftime('%m-%d-%Y')
    filename = 'Unix_Configuration_Report' + timestamp + '.xlsx'
    wb.save(filename)


    #c and f are just arbitrary vars right now for page one and two :\
    c.style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))
    f.style = Style(font=Font(name='Arial', size=11, bold=True, vertAlign=None, color='FF000000'))

    command = 'echo "Unix Configuration Report" | mutt -a "' + filename + '" -s "Unix Configuration Report" -- boomer@wellcare.com'
    os.system(command)







#start execution
if __name__ == '__main__':
    print "Getting server information..."
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dashboard.settings')
    get_server_data()

    #print "Elapsed time: " + str(elapsed_time)







